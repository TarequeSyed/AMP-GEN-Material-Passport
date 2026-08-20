"""Stage 3: write JSON, populate the provided XLSX template and create a chart."""
from pathlib import Path
import json
import re
from collections import Counter
import openpyxl
import matplotlib.pyplot as plt

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "intermediate" / "normalized_items.json"
TEMPLATE = ROOT / "input" / "AMP_Passport_Template.xlsx"
OUT = ROOT / "output"


def find_header_row(ws):
    for row in ws.iter_rows():
        values = [str(c.value or "").strip().lower() for c in row]
        if "gmap id" in values and "description" in values:
            return row[0].row
    raise ValueError("Could not find passport header row")


def export():
    OUT.mkdir(exist_ok=True)
    items = json.loads(DATA.read_text(encoding="utf-8"))
    (OUT / "passport.json").write_text(json.dumps(items, indent=2, ensure_ascii=False), encoding="utf-8")

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["Material Passport"]
    header_row = find_header_row(ws)
    def canon(value):
        return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

    headers = {
        canon(ws.cell(header_row, c).value): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(header_row, c).value
    }
    # Template header names are mapped explicitly so schema changes fail loudly.
    aliases = {
        "gmap_id": ["gmap id"], "boq_item_no": ["boq item no"],
        "description": ["description"], "floor_section": ["floor / section", "floor section"],
        "discipline": ["discipline"], "material_product": ["material / product", "material product"],
        "all_materials_detected": ["all materials detected"], "material_category": ["material category"],
        "material_confidence": ["material confidence"], "grade": ["grade"], "mix_ratio": ["mix ratio"],
        "original_quantity": ["original quantity"], "original_unit": ["original unit"],
        "volume_m3": ["volume m3", "volume (m3)"], "area_m2": ["area m2", "area (m2)"],
        "length_m": ["length m", "length (m)"], "weight_kg": ["weight kg", "weight (kg)"],
        "count_nos": ["count nos", "count (nos)"], "derived_quantity": ["derived quantity"],
        "derived_quantity_unit": ["derived quantity unit"], "schedule": ["schedule"],
        "schedule_item_code": ["schedule item code"], "classification_matched": ["classification matched"],
        "density_kg_m3": ["density kg/m³", "density (kg/m3)", "density (kg/m³)", "density"],
        "embodied_carbon_kg_co2e": ["embodied carbon a1-a3 kg co2e", "embodied carbon a1-a3 (kg co2e)", "embodied carbon a1-a3 (kg co₂e)", "embodied carbon"],
        "gwp_per_kg_co2e": ["gwp / kg (kg co2e/kg)", "gwp / kg (kg co₂e/kg)", "gwp / kg", "gwp"],
        "comment": ["comment", "comments"],
    }
    for offset, item in enumerate(items, start=header_row + 1):
        for key, candidates in aliases.items():
            col = next((headers[canon(x)] for x in candidates if canon(x) in headers), None)
            if col:
                ws.cell(offset, col).value = item.get(key)
    wb.save(OUT / "passport_filled.xlsx")

    # Aggregate data for side-by-side visualization
    categories = sorted(list({x["material_category"] for x in items if x.get("material_category")}))
    cat_counts = {cat: 0 for cat in categories}
    cat_carbon = {cat: 0.0 for cat in categories}

    for x in items:
        cat = x.get("material_category") or "Other"
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        cat_carbon[cat] = cat_carbon.get(cat, 0.0) + float(x.get("embodied_carbon_kg_co2e") or 0.0)

    # Sort data for clean plots
    sorted_by_count = sorted(cat_counts.items(), key=lambda val: val[1], reverse=True)
    labels_count, values_count = zip(*sorted_by_count)

    sorted_by_carbon = sorted(cat_carbon.items(), key=lambda val: val[1], reverse=True)
    labels_carbon, values_carbon = zip(*sorted_by_carbon)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

    # Plot 1: Line Item Counts
    ax1.bar(labels_count, values_count, color="#4F81BD")
    ax1.set_ylabel("Number of BoQ Line Items", fontweight="bold")
    ax1.set_title("Material Category Distribution (by Item Count)", fontweight="bold")
    ax1.tick_params(axis='x', rotation=35)
    for tick in ax1.get_xticklabels():
        tick.set_ha("right")

    # Plot 2: Embodied Carbon in kg CO2e
    ax2.bar(labels_carbon, values_carbon, color="#C0504D")
    ax2.set_ylabel("Embodied Carbon (kg CO₂e)", fontweight="bold")
    ax2.set_title("Material Category Distribution (by Embodied Carbon)", fontweight="bold")
    ax2.tick_params(axis='x', rotation=35)
    for tick in ax2.get_xticklabels():
        tick.set_ha("right")

    plt.suptitle("CBRI Principal's Residence: Material & Carbon Passport Summary", fontsize=14, weight="bold")
    plt.tight_layout()
    plt.savefig(OUT / "visualization.png", dpi=200)
    plt.close()
    print("Outputs written to", OUT)



if __name__ == "__main__":
    export()
